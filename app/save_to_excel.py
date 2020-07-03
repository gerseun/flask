import openpyxl
from openpyxl.styles import Font, Fill, Alignment
import os
from datetime import datetime
from shutil import copy2
import datetime
import pprint
import os.path
from app import set_folder as s     #s verrà usata per richiamare le funzioni in set_folder
from flask import send_file
import urllib.request

def save_xlsx_Produzione(array):
    #OTTENGO UN SINGOLO ARTICOLO
    #prendo le variabili da salvare
    t_imp = array['t_imp'][0]
    t_art = array['t_art'][0]
    cod_art = t_art['cod_art']

    #controllo e creo la cartella
    imp = t_imp["cod_imp"]
    imp_folder = imp.replace("/","-")
    #1 - creo DIR
    s.setFolder(imp_folder)
    #controllo se esiste già -> creo file ARTICOLO-1
    path = 'C:/Produzione Python/'+imp_folder+'/'+cod_art+'.xlsx'
    contPath = 0
    while True:
        if os.path.isfile(path):
            #esiste
            contPath += 1
            path = 'C:/Produzione Python/'+imp_folder+'/'+cod_art+'-'+str(contPath)+'.xlsx'
        else:
            break

    #creo il file excel
    copy2('template taglio.xlsx', path)
    #vado a popolare il file
    popolateFile(array, path)
    return 'file excel modificato'

def popolateFile(insieme, fileName):
    #prendo i dati dell' ordine
    t_imp = insieme['t_imp'][0]
    t_art = insieme['t_art'][0]
    t_comp = insieme['t_comp']
    #popolo INTESTAZIONE -> ARTICOLO ED IMPEGNO
    wb = openpyxl.load_workbook(fileName)

    #creo array pagine e ciclo le pagine!!!!!
    arrayPage = ["TAGLIO", "ORDINE", "MAGAZZINO", "UFF TECNICO"]
    for page in arrayPage:
        ws = wb[page]
        if page != "TAGLIO":
            ws["A1"] = "*A." + str(t_art["id_riga_imp"]) + "*"                  #ID RIGA ART
        ws["B3"] = str(t_imp["cliente"])                                        #CLIENTE ORDINE
        ws["O1"] = str(t_imp["cod_imp"])                                        #CODICE IMPEGNO
        ws["V1"] = str(t_art["data_cons_art"])                                  #DATA CONSEGNA
        ws["AD1"] = str(adesso())                                               #DATA COMPILAZIONE
        ws["O3"] = str(t_art["desc_art"])                                       #DESCRIZIONE ARTICOLO
        ws["AB3"] = str(t_art["cod_art"])                                       #CODICE ARTICOLO
    #fine ciclo header

    #popolo TABELLA -> COMPONENTI
    contT = 6
    contO = 6
    contM = 6
    contUT = 6
    contRow = 6

    for comp in t_comp:
        #INSERISCO LA RIGA SOLO SE QT > 0
        if int(comp["qt_comp"]) > 0:
            #controllo se da tagliare o se da ordinare
            #TAGLIO
            if int(comp["id_produzione"]) == 2:
                #setto il foglio
                ws = wb[arrayPage[0]]
                contT += 2
                contRow = contT
            #ORDINE
            elif int(comp["id_produzione"]) == 1:
                #setto il foglio
                ws = wb[arrayPage[1]]
                contO += 2
                contRow = contO
            #MAGAZZINO
            elif int(comp["id_produzione"]) == 3:
                #setto il foglio
                ws = wb[arrayPage[2]]
                contM += 2
                contRow = contM
            #UFF TECNICO
            elif int(comp["id_produzione"]) == 4:
                #setto il foglio
                ws = wb[arrayPage[3]]
                contUT += 2
                contRow = contUT
            #inserisco la riga componente
            print(ws)
            ws["A" + str(contRow)] = "*C." + str(comp["id_riga_dett"]) + "*"    #ID RIGA COMP
            ws["B" + str(contRow)] = str(comp["qt_comp"])                       #QT COMPO
            ws["D" + str(contRow)] = str(comp["cod_comp"])                      #DIS PARTICOLARE
            ws["K" + str(contRow)] = str(comp["desc_comp"])                     #DESCRIZIONE
            ws["Q" + str(contRow)] = str(comp["dim_comp"])                      #DIMENSIONI
            ws["Y" + str(contRow)] = str(comp["mat_comp"])                      #MATERIALE
            ws["AI" + str(contRow)] = comp["grezzo"]                            #COD GREZZO

    #fine ciclo componenti
    #wb.active = ws["TAGLIO"]
    wb.save(fileName)
    return "OK"

def adesso():
    now = datetime.datetime.now()
    dataOra = now.strftime("%d/%m/%Y")
    return dataOra

def get_cell_coord(wb, range_name):
    my_range = wb.defined_names[range_name]
    dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples
    coord_arr = []
    for title, coord in dests:
        coord_arr.append(coord)
    return coord_arr

#STAMPO IL MATERIALE TAGLIATO
def save_xlsx_Taglio(array):
    t_comp = array["t_compIsorella"]
    #OTTENGO UN INSIEME DI COMPONENTI
    #1 - creo DIR
    s.setFolder("cassone")
    #controllo se esiste già -> creo file cassone
    path = 'C:/Produzione Python/cassone/stampoTaglio.xlsx'
    #creo il file excel
    copy2('template cassone.xlsx', path)
    #vado a popolare il file
    wb = openpyxl.load_workbook(path)
    ws = wb["CASSONE"]
    contRow = 2

    for comp in t_comp:
        ws["A" + str(contRow)] = "*C." + str(comp["id_riga_dett"]) + "*"        #ID RIGA COMP
        ws["B" + str(contRow)] = str(comp["qt_comp"])                           #QT COMPO
        ws["C" + str(contRow)] = str(comp["cod_comp"])                          #DIS PARTICOLARE
        ws["D" + str(contRow)] = str(comp["desc_comp"])                         #DESCRIZIONE
        ws["E" + str(contRow)] = str(comp["dim_comp"])                          #DIMENSIONI
        ws["F" + str(contRow)] = str(comp["mat_comp"])                          #MATERIALE
        ws["F" + str(contRow)] = str(comp["cod_imp"])                           #IMPEGNO
        contRow = contRow + 2

    #fine ciclo componenti
    wb.save(path)

    #scarico il file
    Save_url = 'C:/stampoTaglio.xlsx'
    urllib.request.urlretrieve(path, Save_url)

    return path

    #stampo
    #os.startfile(path,'print')
    #return 'file excel modificato'
